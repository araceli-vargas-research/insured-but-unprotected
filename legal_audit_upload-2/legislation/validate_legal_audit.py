#!/usr/bin/env python3
"""Validate the structure and minimum audit trail of the legal workbook."""

from pathlib import Path
import sys

from openpyxl import load_workbook


WORKBOOK = Path(__file__).with_name("state_health_insurance_legal_audit.xlsx")
JURISDICTIONS = {
    "Alabama", "Alaska", "Arizona", "Arkansas", "California", "Colorado",
    "Connecticut", "Delaware", "District of Columbia", "Florida", "Georgia",
    "Hawaii", "Idaho", "Illinois", "Indiana", "Iowa", "Kansas", "Kentucky",
    "Louisiana", "Maine", "Maryland", "Massachusetts", "Michigan", "Minnesota",
    "Mississippi", "Missouri", "Montana", "Nebraska", "Nevada", "New Hampshire",
    "New Jersey", "New Mexico", "New York", "North Carolina", "North Dakota",
    "Ohio", "Oklahoma", "Oregon", "Pennsylvania", "Rhode Island",
    "South Carolina", "South Dakota", "Tennessee", "Texas", "Utah", "Vermont",
    "Virginia", "Washington", "West Virginia", "Wisconsin", "Wyoming",
}
YEARS = set(range(2015, 2027))


def rows_as_dicts(ws):
    headers = [cell.value for cell in ws[1]]
    for values in ws.iter_rows(min_row=2, values_only=True):
        if any(value is not None for value in values):
            yield dict(zip(headers, values))


def main():
    failures = []
    wb = load_workbook(WORKBOOK, read_only=True, data_only=False)
    required = {"README", "Event Coding", "State-Year Panel", "Source Log", "Codebook", "QA Checklist"}
    missing = required.difference(wb.sheetnames)
    if missing:
        failures.append(f"Missing sheets: {sorted(missing)}")

    events = list(rows_as_dicts(wb["Event Coding"]))
    panel = list(rows_as_dicts(wb["State-Year Panel"]))

    if len(events) != 306:
        failures.append(f"Expected 306 event inventory rows; found {len(events)}")
    if len(panel) != 612:
        failures.append(f"Expected 612 state-year rows; found {len(panel)}")

    event_ids = [r.get("event_id") for r in events]
    if len(event_ids) != len(set(event_ids)):
        failures.append("Duplicate event_id values")
    if {r.get("state_name") for r in events} != JURISDICTIONS:
        failures.append("Event Coding does not contain the exact 51-jurisdiction universe")
    pa_inventory = [r for r in events if r.get("policy_family") == "Prior authorization procedure"]
    if len(pa_inventory) != 51:
        failures.append(f"Expected 51 prior-authorization inventory rows; found {len(pa_inventory)}")
    if any(r.get("review_status") == "Not Started" for r in pa_inventory):
        failures.append("At least one prior-authorization row has not completed secondary-source inventory")

    panel_keys = [(r.get("state_name"), r.get("year")) for r in panel]
    if len(panel_keys) != len(set(panel_keys)):
        failures.append("Duplicate state-year keys")
    if {r.get("state_name") for r in panel} != JURISDICTIONS:
        failures.append("State-Year Panel does not contain the exact 51-jurisdiction universe")
    if {r.get("year") for r in panel} != YEARS:
        failures.append("State-Year Panel must cover 2015 through 2026")

    for row in events:
        if row.get("review_status") == "Verified - Second Review":
            needed = ["legal_citation", "effective_date", "primary_source_url", "reviewer_1", "reviewer_2", "last_verified"]
            absent = [field for field in needed if not row.get(field)]
            if absent:
                failures.append(f"{row.get('event_id')}: second-review verification missing {absent}")

    for row in panel:
        if row.get("construction_status") == "Verified" and not row.get("source_event_ids"):
            failures.append(f"{row.get('state_name')} {row.get('year')}: verified panel row lacks source_event_ids")

    if failures:
        print("LEGAL AUDIT VALIDATION FAILED")
        for failure in failures:
            print(f"- {failure}")
        return 1

    print("LEGAL AUDIT VALIDATION PASSED")
    print(f"- {len(events)} event inventory rows")
    print(f"- {len(panel)} state-year rows")
    print("- 51 jurisdictions; years 2015-2026")
    return 0


if __name__ == "__main__":
    sys.exit(main())
